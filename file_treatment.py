from openpyxl import load_workbook
import csv
import openpyxl
from openpyxl.styles import Font


def change_csv_to_xlsx(name):#изменяет csv yf xlsx
    wb = openpyxl.Workbook()
    ws = wb.active
    with open(name) as f:
        reader = csv.reader(f, delimiter=',')
        for row in reader:
            ws.append(row)
        wb.save(name.replace('csv', 'xlsx'))


def draw(name):#изменяет цвет ячейки по условию
    name = name.replace('csv', 'xlsx')
    wb_file = load_workbook(name)  # , data_only=True)
    ws_file = wb_file.active
    sheet_form = wb_file["Sheet"]
    # sheet_form["A1"].value = "Link"

    k = 0
    for iter, value in enumerate(sheet_form.values):
        if iter == 0:
            continue
        if value[11] != "Не указано" and int(value[11].replace(" ", "")) - int(
                value[7].replace(" ", "")) > 1000000:  # Изменение цены
            k += 1
            a = ws_file["A" + str(iter + 1)]
            a.font = Font(color="FF0000")
    print("Количество объявлений которые подешевели ", k)
    wb_file.save(name)




